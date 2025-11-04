from openpyxl import Workbook, load_workbook
from datetime import timedelta
import config
import os

excel_file = config.LOG_FILE


def init_excel():
    """Initialize Excel file if not exists."""
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Dota2 Log"
        ws.append(["Date", "Played Time", "Did Shut"])
        wb.save(excel_file)


def log_to_excel(date, played_time, did_shut="No"):
    """Log or update today's entry in Excel."""
    init_excel()
    wb = load_workbook(excel_file)
    ws = wb.active

    # Convert seconds → HH:MM:SS
    played_str = str(timedelta(seconds=int(played_time)))

    # 🔍 Check if today's date already exists
    found = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == date:
            ws.cell(row=row, column=2, value=played_str)
            ws.cell(row=row, column=3, value=did_shut)
            found = True
            break

    # If not found → append new row
    if not found:
        ws.append([date, played_str, did_shut])

    wb.save(excel_file)
    print(f"[Excel] Saved: {date}, {played_str}, Shut={did_shut}")


def get_last_played_time(date_str):
    """Return last saved played time for given date in seconds, or 0 if no log."""
    init_excel()
    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        if row[0] == date_str:
            last_played_str = row[1]

            if not last_played_str:
                return 0

            # Case 1: stored as datetime.time
            if hasattr(last_played_str, "hour"):
                return last_played_str.hour * 3600 + last_played_str.minute * 60 + last_played_str.second

            # Case 2: stored as "HH:MM:SS" string
            if isinstance(last_played_str, str):
                h, m, s = map(int, last_played_str.split(":"))
                return h * 3600 + m * 60 + s

            return 0  # fallback

    return 0  # if today's row not found
